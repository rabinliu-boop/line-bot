from flask import Flask, request
import requests
import json
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# =========================
# ⚠️ LINE Token
# =========================
CHANNEL_ACCESS_TOKEN = "Mn03tLk/2u8uW4mUHoS+9Z7xfkvqDpeAhfgPCP4wN/BcaEiNEvkU5WYvBaeZrSYMFiVetB6aOHp0zLpQ46RJ0GSnJ+15+XUrqrZsD15/iHm/MilMrvlbcrpeEm2pMC9/DlWvU3D1FD02E7ea9qYMwQdB04t89/1O/w1cDnyilFU="

# =========================
# Excel 檔案
# =========================
EXCEL_FILE = "data.xlsx"


# =========================
# 初始化 Excel
# =========================
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "工地資料"

        ws.append(["日期", "工地", "進度", "人員"])

        wb.save(EXCEL_FILE)
        print("📁 Excel 已建立")


# =========================
# 首頁
# =========================
@app.route("/")
def home():
    return "Bot is running"


# =========================
# webhook
# =========================
@app.route("/callback", methods=["POST"])
def callback():

    try:
        data = request.get_json(silent=True)

        print("🔥 RAW:", data)

        if not data:
            return "OK"

        for event in data.get("events", []):

            if event.get("type") != "message":
                continue

            text = event["message"].get("text", "")
            reply_token = event.get("replyToken")

            reply = handle_message(text)

            reply_message(reply_token, reply)

        return "OK"

    except Exception as e:
        print("❌ callback error:", str(e))
        return "OK"


# =========================
# 訊息處理
# =========================
def handle_message(text):

    print("📩 RAW:", repr(text))

    # 🔥 全形轉半形
    text = str(text).replace("！", "!").strip()

    print("🧹 CLEAN:", repr(text))

    # =========================
    # 相簿模式
    # =========================
    if text.startswith("!"):

        raw = text[1:].strip()

        result = parse_album(raw)

        print("📊 RESULT:", result)

        if result:

            # 🔥 寫入 Excel
            save_to_excel(result)

            return (
                "📌解析成功\n"
                f"日期: {result['date']}\n"
                f"工地: {result['site']}\n"
                f"進度: {result['progress']}\n"
                f"人員: {result['person']}"
            )

        return (
            "❌格式錯誤\n\n"
            "正確格式：\n"
            "!115.04.15大莊園：外牆拆架完成（銘）"
        )

    return "你說的是：" + text


# =========================
# 解析器（已修正 trim）
# =========================
def parse_album(text):

    try:
        if "：" not in text:
            return None

        left, right = text.split("：", 1)

        date = left[:10].strip()     # ✔ 修正
        site = left[10:].strip()     # ✔ 修正

        if "（" not in right or "）" not in right:
            return None

        progress = right.split("（")[0].strip()
        person = right.split("（")[1].replace("）", "").strip()

        return {
            "date": date,
            "site": site,
            "progress": progress,
            "person": person
        }

    except Exception as e:
        print("❌ parse error:", str(e))
        return None


# =========================
# Excel 寫入（穩定版）
# =========================
def save_to_excel(data):

    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        ws.append([
            data["date"].strip(),
            data["site"].strip(),
            data["progress"].strip(),
            data["person"].strip()
        ])

        wb.save(EXCEL_FILE)

        print("💾 已寫入 Excel成功：", data)

    except Exception as e:
        print("❌ Excel error:", str(e))


# =========================
# LINE reply（穩定版）
# =========================
def reply_message(reply_token, text):

    try:
        url = "https://api.line.me/v2/bot/message/reply"

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + CHANNEL_ACCESS_TOKEN.strip()
        }

        payload = {
            "replyToken": reply_token,
            "messages": [
                {"type": "text", "text": text}
            ]
        }

        res = requests.post(
            url,
            headers=headers,
            json=payload,
            timeout=5
        )

        print("📤 status:", res.status_code)
        print("📤 response:", res.text)

    except Exception as e:
        print("❌ reply error:", str(e))


# =========================
# 啟動初始化
# =========================
init_excel()


if __name__ == "__main__":
    app.run()
